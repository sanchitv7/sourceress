# sourceress/src/sourceress/agents/excel_writer.py

"""Excel Writer Agent.

Emits all artefacts into a single Excel workbook.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from sourceress.models import PitchResult

GREEN_GRADIENT_FILL = PatternFill(start_color="C6EFCE", end_color="006100", fill_type="solid")


class ExcelWriter:
    """Agent responsible for writing outputs to an Excel (.xlsx) file."""

    name: str = "excel_writer"

    async def run(
        self,
        pitched: PitchResult,
        output_path: Path | str = "output.xlsx",
        **kwargs: Any,
    ) -> Path:  # noqa: D401
        """Execute the agent.

        Args:
            pitched: Final artefacts from PitchGenerator.
            output_path: Destination file path.
            **kwargs: Additional runtime parameters.

        Returns:
            Path to the written Excel file.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "candidates"
        logger.debug("Writing %d candidate rows to Excel", len(pitched.pitches))

        # Header
        headers = [
            "Candidate Name",
            "LinkedIn URL",
            "Match Score",
            "Key Matches",
            "Pitch Script",
            "LinkedIn DM",
            "WhatsApp Msg",
            "Notes",
        ]
        ws.append(headers)
        ws.freeze_panes = "A2"

        # TODO: Write real data rows
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].auto_size = True  # type: ignore[attr-defined]

        match_score_col = headers.index("Match Score") + 1
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=match_score_col)
            cell.fill = GREEN_GRADIENT_FILL

        wb.save(output_path)
        logger.debug("Excel workbook saved to %s", output_path)
        return Path(output_path) 